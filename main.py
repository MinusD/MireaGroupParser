import openpyxl

INPUT_FILE_NAME = 'Vybor_profilya_dlya_publikatsii.xlsx'  # Таблица с распределением
OUTPUT_FILE_NAME = 'groups.txt'  # Выходной файл со списками
NAME_FORMAT = '{}. {} ({})\n'  # Формат имени в списке, (номер, ФИО, бывшая группа)
GROUP_FORMAT = '\n\nГруппа {}\n'  # Формат названия группы


class Parser:
    def __init__(self) -> None:
        self.data: list = []
        self.groups: set = set()

    def _parse(self) -> None:
        """Парсит файл с распределением"""
        book = openpyxl.load_workbook(INPUT_FILE_NAME)
        sheet = book.active
        num_rows = sheet.max_row
        for i in range(2, num_rows):
            self.data.append([sheet.cell(column=1, row=i).value, sheet.cell(column=2, row=i).value,
                              sheet.cell(column=3, row=i).value])
            self.groups.add(sheet.cell(column=3, row=i).value)

    def generate_file(self, filename: str = OUTPUT_FILE_NAME) -> None:
        """Генерируем файл с группами"""
        with open(OUTPUT_FILE_NAME, 'w') as f:
            self._parse()
            groups = sorted(list(self.groups))
            for group in groups:
                group_list_tmp = []
                for student in self.data:
                    if student[2] == group:
                        group_list_tmp.append(student[:2])
                group_list_tmp.sort(key=lambda x: x[0])
                f.write(GROUP_FORMAT.format(group))
                f.writelines([NAME_FORMAT.format(i + 1, group_list_tmp[i][0], group_list_tmp[i][1]) for i in
                              range(len(group_list_tmp))])


if __name__ == '__main__':
    parser = Parser()
    parser.generate_file()
